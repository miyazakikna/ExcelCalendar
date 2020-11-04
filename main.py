# -*- coding: utf-8 -*-
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.styles import Alignment
from openpyxl.styles.borders import Border, Side
import calendar
import datetime
import jpholiday


def main():
    # ワークブックを新規作成する
    book = openpyxl.Workbook()

    # シートを取得し名前を変更する
    sheet = book.active
    sheet.title = 'カレンダー'

    # 現在の日付を取得
    dt_now = datetime.datetime.now()
    year = dt_now.year
    month = dt_now.month
    last_day = calendar.monthrange(year, month)[1]

    # 対象月の日付を書き込む
    date_of_month(sheet, last_day)

    # 対象月の日付に結びつく曜日を書き込み、休日と祝日の色を変更
    date_of_week_jp(sheet, year, month, last_day)

    # 罫線を引く
    set_border(sheet, last_day)

    # ワークブックに名前をつけて保存する
    book.save('カレンダー.xlsx')


# 範囲を指定してセルを取得する
def get_cells(sheet, start_cell, end_cell):
    return sheet[start_cell:end_cell]


# 対象月の日付を書き込む
def date_of_month(sheet, last_day):
    # 書き込む範囲を指定
    start_cell = 'A1'
    end_cell = 'A' + str(last_day)
    select_cells = get_cells(sheet, start_cell, end_cell)
    i = 1
    for row in select_cells:
        for cell in row:
            # 日付を書き込む
            cell.value = i
            i += 1


# 対象月の日付に結びつく曜日を書き込み、休日と祝日の色を変更
def date_of_week_jp(sheet, year, month, last_day):
    day_count = 1
    # 月初の曜日を取得
    dt = datetime.datetime(year, month, 1)
    weekday_index = dt.weekday()
    w_list = ['月', '火', '水', '木', '金', '土', '日']

    # 書き込む範囲を指定
    start_cell = 'B1'
    end_cell = 'B' + str(last_day)
    select_cells = get_cells(sheet, start_cell, end_cell)
    for row in select_cells:
        for cell in row:
            # 日曜日まできたら月曜日に戻す
            if weekday_index == len(w_list):
                weekday_index = 0
            # 曜日を書き込む
            cell.value = w_list[weekday_index]
            # 日付を取得
            date = datetime.date((int(year)), int(month), day_count)
            # 休日と祝日の時、セルの色を変更する
            if w_list[weekday_index] == '土' or w_list[weekday_index] == '日' or jpholiday.is_holiday(date):
                # セルの色を変更する
                holiday_color_cell(sheet, day_count)
            weekday_index += 1
            day_count += 1


# セルの色を変更する
def holiday_color_cell(sheet, row):
    # 背景色を変更範囲指定
    start_cell = 'A' + str(row)
    end_cell = 'C' + str(row)
    select_cell = get_cells(sheet, start_cell, end_cell)
    for rows in select_cell:
        for cell in rows:
            # 背景色を変更
            cell.fill = PatternFill(patternType='solid', fgColor='4169e1')


# 罫線を引く
def set_border(sheet, last_day):
    side = Side(style='thin', color='000000')
    border = Border(top=side, bottom=side, left=side, right=side)
    # 罫線を引く範囲指定
    start_cell = 'A1'
    end_cell = 'C' + str(last_day)
    select_cells = get_cells(sheet, start_cell, end_cell)
    for row in select_cells:
        for cell in row:
            # 罫線を引く
            sheet[cell.coordinate].border = border


if __name__ == "__main__":
    main()
